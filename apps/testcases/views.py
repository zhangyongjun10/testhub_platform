from rest_framework import generics, permissions, status, pagination
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.db import models
import uuid
import traceback
from .models import TestCase, TestCaseStep, TestCaseAttachment, TestCaseComment, TestCaseImportTask
from .serializers import (
    TestCaseSerializer, TestCaseListSerializer, TestCaseCreateSerializer, TestCaseUpdateSerializer,
    TestCaseImportTaskSerializer, TestCaseImportTaskCreateSerializer
)
from apps.projects.models import Project

class TestCasePagination(pagination.PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class TestCaseListCreateView(generics.ListCreateAPIView):
    queryset = TestCase.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = TestCasePagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['priority', 'test_type', 'project']
    search_fields = ['title', 'description']
    ordering_fields = ['created_at', 'updated_at', 'priority']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return TestCaseCreateSerializer
        return TestCaseListSerializer
    
    def get_queryset(self):
        user = self.request.user
        accessible_projects = Project.objects.filter(
            models.Q(owner=user) | models.Q(members=user)
        ).distinct()
        return TestCase.objects.filter(
            project__in=accessible_projects
        ).select_related(
            'author', 'assignee', 'project'
        ).prefetch_related(
            'versions'
        ).distinct()
    
    def get_user_accessible_projects(self, user):
        """获取用户有权限访问的项目"""
        return Project.objects.filter(
            models.Q(owner=user) | models.Q(members=user)
        ).distinct()
    
    def perform_create(self, serializer):
        user = self.request.user
        project_id = self.request.data.get('project_id')
        
        # 获取用户有权限的项目
        accessible_projects = self.get_user_accessible_projects(user)
        
        if project_id:
            # 检查指定的项目是否存在且用户有权限
            try:
                project = accessible_projects.get(id=project_id)
            except Project.DoesNotExist:
                # 如果指定项目不存在或无权限，使用第一个可访问的项目
                project = accessible_projects.first()
                if not project:
                    # 如果用户没有任何项目，创建默认项目
                    project = Project.objects.create(
                        name="默认项目",
                        owner=user,
                        description='系统自动创建的默认项目'
                    )
        else:
            # 没有指定项目，使用第一个可访问的项目
            project = accessible_projects.first()
            if not project:
                # 如果用户没有任何项目，创建默认项目
                project = Project.objects.create(
                    name="默认项目",
                    owner=user,
                    description='系统自动创建的默认项目'
                )
        
        serializer.save(author=user, project=project)

class TestCaseDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = TestCase.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return TestCaseUpdateSerializer
        return TestCaseSerializer
    
    def get_queryset(self):
        user = self.request.user
        accessible_projects = Project.objects.filter(
            models.Q(owner=user) | models.Q(members=user)
        ).distinct()
        return TestCase.objects.filter(
            project__in=accessible_projects
        ).select_related(
            'author', 'assignee', 'project'
        ).prefetch_related(
            'versions', 'step_details', 'attachments', 'comments'
        )
    
    def get_user_accessible_projects(self, user):
        """获取用户有权限访问的项目"""
        return Project.objects.filter(
            models.Q(owner=user) | models.Q(members=user)
        ).distinct()
    
    def perform_update(self, serializer):
        user = self.request.user
        project_id = self.request.data.get('project_id')
        
        if project_id:
            # 检查指定的项目是否存在且用户有权限
            accessible_projects = self.get_user_accessible_projects(user)
            try:
                project = accessible_projects.get(id=project_id)
                serializer.save(project=project)
            except Project.DoesNotExist:
                # 如果指定项目不存在或无权限，保持原项目不变
                serializer.save()
        else:
            # 没有指定项目，保持原项目不变
            serializer.save()

class TestCaseImportTaskListCreateView(generics.ListCreateAPIView):
    """导入任务列表和创建"""
    queryset = TestCaseImportTask.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = TestCasePagination
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['status', 'project']
    ordering_fields = ['created_at', 'updated_at']
    ordering = ['-created_at']
    parser_classes = [MultiPartParser, FormParser]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return TestCaseImportTaskCreateSerializer
        return TestCaseImportTaskSerializer

    def get_queryset(self):
        user = self.request.user
        accessible_projects = Project.objects.filter(
            models.Q(owner=user) | models.Q(members=user)
        ).distinct()

        # 管理员可以看到所有任务，普通用户只能看到自己创建的或相关项目的任务
        if user.is_superuser:
            return TestCaseImportTask.objects.all()
        else:
            return TestCaseImportTask.objects.filter(
                models.Q(created_by=user) | models.Q(project__in=accessible_projects)
            ).distinct().select_related('created_by', 'project')

    def perform_create(self, serializer):
        import excel_import  # 导入Celery任务模块
        user = self.request.user
        file = self.request.FILES.get('file')
        project_id = self.request.data.get('project_id')

        if not file:
            raise serializers.ValidationError({"file": "请上传文件"})

        # 获取项目
        if project_id:
            accessible_projects = Project.objects.filter(
                models.Q(owner=user) | models.Q(members=user)
            ).distinct()
            try:
                project = accessible_projects.get(id=project_id)
            except Project.DoesNotExist:
                raise serializers.ValidationError({"project_id": "无权访问该项目"})
        else:
            # 获取用户的第一个项目
            project = Project.objects.filter(
                models.Q(owner=user) | models.Q(members=user)
            ).first()

        # 生成任务ID
        task_id = str(uuid.uuid4())

        # 保存上传的文件到临时目录
        import os
        from django.conf import settings

        upload_dir = os.path.join(settings.MEDIA_ROOT, 'temp_imports')
        os.makedirs(upload_dir, exist_ok=True)

        file_path = os.path.join(upload_dir, f'{task_id}_{file.name}')
        with open(file_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        # 创建导入任务记录
        import_task = serializer.save(
            task_id=task_id,
            file_name=file.name,
            created_by=user,
            project=project
        )

        # 触发异步导入任务
        excel_import.import_test_cases_from_excel.delay(task_id, file_path, user.id, project.id if project else None)

        return import_task

class TestCaseImportTaskDetailView(generics.RetrieveAPIView):
    """导入任务详情"""
    queryset = TestCaseImportTask.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = TestCaseImportTaskSerializer
    lookup_field = 'task_id'

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def download_import_template(request):
    """下载导入模板"""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, Alignment, PatternFill

    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例导入模板"

    # 设置表头
    headers = [
        '用例标题*', '前置条件', '操作步骤', '预期结果*',
        '优先级', '测试类型', '标签', '关联版本(用逗号分隔)'
    ]

    # 设置表头样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 写入表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 添加示例数据
    example_data = [
        ['用户登录功能测试', '已注册用户账号', '1. 打开登录页面\n2. 输入用户名和密码\n3. 点击登录按钮', '成功跳转到首页', '高', '功能测试', '登录,认证', 'v1.0,v1.1'],
        ['用户注册功能测试', '打开注册页面', '1. 输入用户名\n2. 输入密码\n3. 输入确认密码\n4. 点击注册按钮', '注册成功并跳转到登录页', '中', '功能测试', '注册', 'v1.0'],
    ]

    for row_num, row_data in enumerate(example_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # 设置列宽
    column_widths = [30, 25, 40, 30, 12, 15, 20, 25]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    # 设置行高
    ws.row_dimensions[1].height = 30

    # 准备响应
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=testcase_import_template.xlsx'

    wb.save(response)
    return response
