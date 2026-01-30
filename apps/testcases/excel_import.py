import os
import traceback
from celery import shared_task
from django.core.files.base import File
import openpyxl
from .models import TestCase, TestCaseImportTask
from apps.users.models import User
from apps.projects.models import Project
from apps.versions.models import Version


@shared_task(bind=True, max_retries=3)
def import_test_cases_from_excel(self, task_id, file_path, user_id, project_id):
    """
    异步导入测试用例的Celery任务
    :param task_id: 任务ID
    :param file_path: Excel文件路径
    :param user_id: 用户ID
    :param project_id: 项目ID
    """
    try:
        # 更新任务状态为处理中
        import_task = TestCaseImportTask.objects.get(task_id=task_id)
        import_task.status = 'processing'
        import_task.save()

        # 获取用户和项目
        user = User.objects.get(id=user_id)
        project = Project.objects.get(id=project_id) if project_id else None

        # 读取Excel文件
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # 解析表头，建立列名映射
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row[0] and '标题' in str(row[0]):
                header_row = row_idx
                break

        if not header_row:
            raise ValueError("未找到有效的表头行，请检查Excel文件格式")

        # 读取表头
        headers = [cell for cell in ws[header_row]]
        header_map = {str(cell.value).strip(): col_idx for col_idx, cell in enumerate(headers)}

        # 智能列名映射
        column_mapping = {
            'title': ['用例标题', '标题', '测试用例', 'case title', 'title'],
            'preconditions': ['前置条件', '前提条件', 'preconditions'],
            'steps': ['操作步骤', '步骤', '测试步骤', 'steps'],
            'expected_result': ['预期结果', '期望结果', 'expected result'],
            'priority': ['优先级', 'priority'],
            'test_type': ['测试类型', '用例类型', 'test type'],
            'tags': ['标签', 'tag'],
            'versions': ['关联版本', '版本', 'versions']
        }

        # 找到每列对应的索引
        col_indices = {}
        for field, possible_names in column_mapping.items():
            for name in possible_names:
                if name in [str(h).strip().lower() for h in header_map.keys()]:
                    # 找到匹配的列
                    for header_name, col_idx in header_map.items():
                        if str(header_name).strip().lower() == name.lower():
                            col_indices[field] = col_idx
                            break
                    break

        # 读取数据行
        data_rows = list(ws.iter_rows(min_row=header_row + 1, values_only=True))
        total_count = len(data_rows)

        import_task.total_count = total_count
        import_task.save()

        success_count = 0
        failed_count = 0
        error_details = []

        # 优先级映射
        priority_map = {
            '低': 'low', 'low': 'low',
            '中': 'medium', 'medium': 'medium',
            '高': 'high', 'high': 'high',
            '紧急': 'critical', 'critical': 'critical'
        }

        # 测试类型映射
        test_type_map = {
            '功能测试': 'functional', 'functional': 'functional',
            '集成测试': 'integration', 'integration': 'integration',
            'api测试': 'api', 'api': 'api',
            'ui测试': 'ui', 'ui': 'ui',
            '性能测试': 'performance', 'performance': 'performance',
            '安全测试': 'security', 'security': 'security'
        }

        # 批量导入
        for row_idx, row in enumerate(data_rows, 1):
            try:
                # 更新进度
                progress = int((row_idx - 1) / total_count * 100)
                import_task.progress = progress
                import_task.save()

                # 跳过空行
                if not any(row):
                    continue

                # 提取字段数据
                title = str(row[col_indices.get('title', 0)]).strip() if 'title' in col_indices else ''
                preconditions = str(row[col_indices.get('preconditions', 1)]).strip() if 'preconditions' in col_indices else ''
                steps = str(row[col_indices.get('steps', 2)]).strip() if 'steps' in col_indices else ''
                expected_result = str(row[col_indices.get('expected_result', 3)]).strip() if 'expected_result' in col_indices else ''
                priority_val = str(row[col_indices.get('priority', 4)]).strip() if 'priority' in col_indices else 'medium'
                test_type_val = str(row[col_indices.get('test_type', 5)]).strip() if 'test_type' in col_indices else 'functional'
                tags_val = str(row[col_indices.get('tags', 6)]).strip() if 'tags' in col_indices else ''
                versions_val = str(row[col_indices.get('versions', 7)]).strip() if 'versions' in col_indices else ''

                # 验证必填字段
                if not title or title == 'None':
                    raise ValueError("用例标题不能为空")

                if not expected_result or expected_result == 'None':
                    raise ValueError("预期结果不能为空")

                # 处理优先级
                priority = priority_map.get(priority_val.lower(), 'medium')

                # 处理测试类型
                test_type = test_type_map.get(test_type_val.lower(), 'functional')

                # 处理标签
                tags = []
                if tags_val and tags_val != 'None':
                    tags = [tag.strip() for tag in tags_val.split(',') if tag.strip()]

                # 处理版本关联
                versions = []
                if versions_val and versions_val != 'None':
                    version_names = [v.strip() for v in versions_val.split(',') if v.strip()]
                    for version_name in version_names:
                        try:
                            version = Version.objects.filter(
                                project=project,
                                name__icontains=version_name
                            ).first()
                            if version:
                                versions.append(version)
                        except Exception as e:
                            pass  # 忽略版本关联错误

                # 转换换行符为<br>标签
                def convert_newline_to_br(text):
                    if not text or text == 'None':
                        return ''
                    return text.replace('\n', '<br>')

                # 创建测试用例
                testcase = TestCase.objects.create(
                    project=project,
                    title=title,
                    description='',
                    preconditions=convert_newline_to_br(preconditions),
                    steps=convert_newline_to_br(steps),
                    expected_result=convert_newline_to_br(expected_result),
                    priority=priority,
                    test_type=test_type,
                    tags=tags,
                    author=user
                )

                # 关联版本
                if versions:
                    testcase.versions.set(versions)

                success_count += 1

            except Exception as e:
                failed_count += 1
                error_msg = f"第{row_idx}行导入失败: {str(e)}"
                error_details.append({
                    'row': row_idx,
                    'error': str(e),
                    'data': {str(headers[i]): str(row[i]) if i < len(row) else '' for i in range(min(len(headers), len(row)))}
                })
                continue

        # 删除临时文件
        try:
            os.remove(file_path)
        except:
            pass

        # 更新任务状态
        import_task.progress = 100
        import_task.success_count = success_count
        import_task.failed_count = failed_count
        import_task.error_details = error_details

        if failed_count == 0:
            import_task.status = 'success'
        elif success_count == 0:
            import_task.status = 'failed'
        else:
            import_task.status = 'partial_success'

        import_task.save()

        return {
            'task_id': task_id,
            'status': import_task.status,
            'total_count': total_count,
            'success_count': success_count,
            'failed_count': failed_count
        }

    except Exception as e:
        # 删除临时文件
        try:
            os.remove(file_path)
        except:
            pass

        # 更新任务状态为失败
        try:
            import_task = TestCaseImportTask.objects.get(task_id=task_id)
            import_task.status = 'failed'
            import_task.error_details = [{'error': f"任务执行失败: {str(e)}", 'traceback': traceback.format_exc()}]
            import_task.save()
        except:
            pass

        # 重试逻辑
        if self.request.retries < self.max_retries:
            raise self.retry(exc=e, countdown=60 * (self.request.retries + 1))

        raise e
